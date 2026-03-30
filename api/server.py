"""
api/server.py — FastAPI SSE 流式分析后端

端点列表:
  POST /api/v1/analyze          → Server-Sent Events (SSE) 流式分析
  POST /api/v1/health-check     → 持仓体检（JSON 请求/响应）
  POST /api/v1/chat             → 财商学长 AI 对话（Qwen）
  POST /api/v1/trade/order      → 模拟撮合下单
  GET  /api/v1/market/quotes    → 实时行情列表
  GET  /api/v1/portfolio/summary → 虚拟账户持仓摘要
  GET  /api/v1/health           → 健康检查
  GET  /api/v1/graph/mermaid    → 图拓扑 Mermaid 字符串

SSE 事件格式:
  每个事件结构为 JSON:
  {
    "event":     "<event_type>",    // 见下方事件类型说明
    "node":      "<node_name>",     // 产生此事件的节点名称
    "message":   "<human_msg>",     // 人类可读摘要
    "data":      { ... },           // 节点输出的结构化数据
    "timestamp": "<ISO8601>",       // 事件时间戳
    "seq":       <int>              // 事件序号（从1开始）
  }

事件类型 (event):
  node_start   — 节点开始执行（含 node 名称）
  node_complete— 节点执行完成（含 data 输出）
  conflict     — 检测到基本面/技术面冲突，触发辩论
  debate       — 辩论完成
  risk_check   — 风控审批结果
  risk_retry   — 风控拒绝，要求修订
  trade_order  — 最终交易指令生成
  complete     — 全流程完成
  error        — 发生错误

启动命令:
  uvicorn api.server:app --host 0.0.0.0 --port 8000 --reload
"""
from __future__ import annotations

import asyncio
import json
import random
import uuid
from datetime import datetime, timedelta, timezone
from typing import AsyncGenerator, Optional

from fastapi import Depends, FastAPI, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.concurrency import run_in_threadpool
from loguru import logger
from pydantic import BaseModel, Field, field_validator
from utils.market_classifier import MarketClassifier
from config import config as _cfg

# ── DB & Auth 延迟导入（避免启动前引用未初始化的 engine）─────
def _get_db_dep():
    from db.engine import get_db
    return get_db()

# 实际 Depends 需要 callable，用 lambda 包装
from db.engine import get_db as _db_get
_get_db_dep = _db_get

# 国内金融数据域名 + DashScope LLM 域名加入 NO_PROXY
# 防止 TUN/VPN 代理截断对东方财富、新浪等数据源的直连请求（RemoteDisconnected 根因）
import os as _os
_no_proxy_extra = (
    "dashscope.aliyuncs.com,aliyuncs.com,"
    "eastmoney.com,push2.eastmoney.com,push2his.eastmoney.com,"
    "datacenter-web.eastmoney.com,np-anotice-stock.eastmoney.com,"
    "hq.sinajs.cn,sinajs.cn,sina.com.cn,finance.sina.com.cn,"
    "money.126.net,netease.com,10jqka.com.cn,xueqiu.com,"
    "akshare.xyz"
)
_existing_no_proxy = _os.environ.get("NO_PROXY", _os.environ.get("no_proxy", ""))
_merged_no_proxy = (
    _existing_no_proxy + "," + _no_proxy_extra if _existing_no_proxy else _no_proxy_extra
)
_os.environ["NO_PROXY"] = _merged_no_proxy
_os.environ["no_proxy"] = _merged_no_proxy

# ════════════════════════════════════════════════════════════════
# 应用初始化
# ════════════════════════════════════════════════════════════════

app = FastAPI(
    title="Multi-Agent Trading System API",
    description="LangGraph + FAISS RAG 驱动的量化交易多智能体分析服务",
    version="2.0.0",
    docs_url="/docs",
    redoc_url="/redoc",
)

# CORS（允许 Streamlit / Vue / React 前端跨域调用）
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ════════════════════════════════════════════════════════════════
# 启动时初始化（FAISS 知识库 + LangGraph 图）
# ════════════════════════════════════════════════════════════════

_compiled_graph = None

# ════════════════════════════════════════════════════════════════
# 全局市场数据缓存（Background Polling，TTL = 5 分钟）
# ════════════════════════════════════════════════════════════════
_MARKET_CACHE: dict = {"indices": [], "news": [], "sectors": [], "sentiment": {}, "ts": 0.0}


async def _market_data_poller():
    """后台循环：启动后立即预热，之后每 5 分钟刷新大盘指数和市场快讯缓存。
    若首次预热后缓存仍有缺失，30s 后自动重试，直到数据就绪。"""
    logger.info("[market_poller] 后台数据预热任务启动")
    while True:
        try:
            from tools.market_data import (
                get_market_indices_raw, get_market_news_raw,
                get_sector_data_raw, get_market_sentiment_raw,
            )
            loop = asyncio.get_event_loop()
            indices, news, sectors, sentiment = await asyncio.gather(
                loop.run_in_executor(None, get_market_indices_raw),
                loop.run_in_executor(None, get_market_news_raw, 20),
                loop.run_in_executor(None, get_sector_data_raw),
                loop.run_in_executor(None, get_market_sentiment_raw),
                return_exceptions=True,
            )
            if not isinstance(indices, Exception) and indices:
                _MARKET_CACHE["indices"] = indices
            if not isinstance(news, Exception) and news:
                _MARKET_CACHE["news"] = news
            if not isinstance(sectors, Exception) and sectors:
                _MARKET_CACHE["sectors"] = sectors
            if not isinstance(sentiment, Exception) and sentiment:
                _MARKET_CACHE["sentiment"] = sentiment
            _MARKET_CACHE["ts"] = datetime.now(timezone.utc).timestamp()

            fallback_n = sum(1 for r in _MARKET_CACHE["indices"] if r.get("is_fallback"))
            logger.info(
                f"[market_poller] 缓存刷新完成: "
                f"{len(_MARKET_CACHE['indices'])} 个指数(fallback={fallback_n}), "
                f"{len(_MARKET_CACHE['news'])} 条快讯"
            )
            # 若仍有 fallback 数据，30s 后重试（否则等满 5 分钟）
            sleep_s = 30 if fallback_n > 0 else 300
        except Exception as e:
            logger.warning(f"[market_poller] 刷新异常: {e}")
            sleep_s = 30   # 出错也快速重试
        await asyncio.sleep(sleep_s)


@app.on_event("startup")
async def startup_event():
    """应用启动时预热 FAISS 知识库和 LangGraph 图"""
    global _compiled_graph
    logger.info("🚀 Trading System API 启动中...")

    # 0. 初始化数据库（建表，幂等）
    try:
        from db.engine import init_db
        await init_db()
        logger.info("✅ 数据库初始化完成（campusquant.db）")
    except Exception as e:
        logger.error(f"❌ 数据库初始化失败: {e}")

    # 1. 初始化 FAISS 知识库
    try:
        from tools.knowledge_base import init_knowledge_base
        ok = init_knowledge_base()
        logger.info(f"{'✅' if ok else '⚠️'} FAISS 知识库初始化: {'成功' if ok else '降级模式'}")
    except Exception as e:
        logger.error(f"❌ 知识库初始化失败: {e}")

    # 2. 构建 LangGraph 图（带 MemorySaver）
    try:
        from graph.builder import build_graph_with_memory
        _compiled_graph = build_graph_with_memory()
        logger.info("✅ LangGraph 图构建完成")
    except Exception as e:
        logger.error(f"❌ 图构建失败: {e}")

    # 3. 冷启动热榜预热（后台线程，不阻塞 uvicorn 启动）
    try:
        from tools.hot_news import refresh_in_background
        refresh_in_background()
        logger.info("✅ 热榜缓存预热任务已启动（后台）")
    except Exception as e:
        logger.warning(f"⚠️ 热榜缓存预热失败（非致命）: {e}")

    # 4. 大盘指数 & 市场快讯后台轮询（asyncio 异步任务，每5分钟刷新）
    asyncio.create_task(_market_data_poller())
    logger.info("✅ 市场数据后台轮询已启动（首次预热中...）")

    # 5. 港股 Board Lot 缓存预热（后台，不阻塞启动）
    asyncio.create_task(_load_hk_lot_cache())

    logger.info("✅ API 服务启动完成，监听请求...")


# ════════════════════════════════════════════════════════════════
# 请求/响应模型
# ════════════════════════════════════════════════════════════════

class AnalyzeRequest(BaseModel):
    symbol: str = Field(..., description="交易标的代码 (如 AAPL / 600519.SH)")
    days:   int = Field(default=180, ge=30, le=365, description="历史数据天数")


class HealthResponse(BaseModel):
    status:      str
    version:     str
    graph_ready: bool
    kb_ready:    bool
    timestamp:   str


class PositionItem(BaseModel):
    """持仓体检单条持仓（供 /api/v1/health-check 请求使用）"""
    symbol:   str   = Field(..., description="标的代码，如 600519.SH / AAPL")
    quantity: float = Field(..., ge=0, description="持仓数量（股/份）")
    avg_cost: float = Field(..., ge=0, description="平均持仓成本价")


class PortfolioCheckRequest(BaseModel):
    """POST /api/v1/health-check 请求体"""
    positions: list[PositionItem] = Field(..., min_length=1, description="持仓列表")


class ChatRequest(BaseModel):
    """POST /api/v1/chat 请求体"""
    message:     str            = Field(..., min_length=1, description="用户消息")
    session_key: Optional[str]  = Field(default=None, description="对话 Session UUID（localStorage 存储，持久记忆）")
    history:     list[dict]     = Field(default_factory=list, description="历史对话（兼容旧版，有 session_key 时忽略）")


class TradeOrderRequest(BaseModel):
    """POST /api/v1/trade/order 请求体"""
    symbol:   str   = Field(..., description="标的代码，如 600519.SH / AAPL")
    action:   str   = Field(..., pattern="^(BUY|SELL)$", description="BUY 或 SELL")
    quantity: float = Field(..., gt=0, description="数量（股/份）")


# ════════════════════════════════════════════════════════════════
# SSE 事件构建工具
# ════════════════════════════════════════════════════════════════

def _make_sse_event(
    event:    str,
    node:     str,
    message:  str,
    data:     Optional[dict] = None,
    seq:      int = 0,
) -> str:
    """
    将事件信息序列化为 SSE 格式字符串。

    SSE 协议格式:
      event: <event_type>\\n
      data: <JSON>\\n
      \\n
    """
    payload = {
        "event":     event,
        "node":      node,
        "message":   message,
        "data":      data or {},
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "seq":       seq,
    }
    json_str = json.dumps(payload, ensure_ascii=False)
    return f"event: {event}\ndata: {json_str}\n\n"


# ════════════════════════════════════════════════════════════════
# 核心：LangGraph astream_events → SSE 生成器
# ════════════════════════════════════════════════════════════════

# 节点名称 → 人类可读标签映射
_NODE_LABELS = {
    "data_node":        "数据情报员",
    "rag_node":         "RAG 知识检索",
    "fundamental_node": "基本面分析师",
    "technical_node":   "技术分析师",
    "sentiment_node":   "舆情分析师",
    "portfolio_node":   "基金经理",
    "debate_node":      "多空辩论裁决",
    "risk_node":        "风险控制官",
    "trade_executor":   "交易指令生成",
}


async def _stream_graph_events(
    symbol:    str,
    thread_id: str,
) -> AsyncGenerator[str, None]:
    """
    调用 LangGraph .astream_events()，将图节点事件转译为 SSE 数据流。

    LangGraph astream_events API 返回的事件类型:
      - on_chain_start  : 图/节点开始
      - on_chain_end    : 图/节点结束（含输出）
      - on_chain_stream : 流式输出中间值

    我们只关注各节点（Node）的 start / end 事件。
    """
    from graph.builder import make_initial_state

    initial_state = make_initial_state(symbol)
    config        = {"configurable": {"thread_id": thread_id}}
    seq           = 0

    # ── 发送开始事件 ──────────────────────────────────────────
    seq += 1
    yield _make_sse_event(
        event="start",
        node="system",
        message=f"开始分析 {symbol}，LangGraph 多智能体引擎启动...",
        data={"symbol": symbol, "thread_id": thread_id},
        seq=seq,
    )
    await asyncio.sleep(0)   # 让事件循环有机会 flush

    if _compiled_graph is None:
        seq += 1
        yield _make_sse_event(
            event="error",
            node="system",
            message="LangGraph 图尚未初始化，请稍后重试",
            seq=seq,
        )
        return

    # ── 追踪已通知的节点（避免重复 node_start）────────────────
    notified_start: set[str] = set()
    last_state:     dict     = {}
    _graph_error:   str | None = None  # 记录图执行异常，用于降级研报

    try:
        async for event in _compiled_graph.astream_events(
            initial_state,
            config=config,
            version="v2",            # 使用 LangGraph astream_events v2
        ):
            kind      = event.get("event", "")
            node_name = event.get("name", "")
            run_id    = event.get("run_id", "")

            # ── 节点开始 ──────────────────────────────────────
            if kind == "on_chain_start" and node_name in _NODE_LABELS:
                if node_name not in notified_start:
                    notified_start.add(node_name)
                    label = _NODE_LABELS[node_name]
                    seq += 1
                    yield _make_sse_event(
                        event="node_start",
                        node=node_name,
                        message=f"⚙️ {label} 开始工作...",
                        data={"label": label},
                        seq=seq,
                    )
                    await asyncio.sleep(0)

            # ── 节点完成 ──────────────────────────────────────
            elif kind == "on_chain_end" and node_name in _NODE_LABELS:
                output    = event.get("data", {}).get("output") or {}
                label     = _NODE_LABELS[node_name]

                # 更新最新 state 快照（合并节点输出），用 .get() 防止非 dict output
                if isinstance(output, dict):
                    last_state.update(output)
                    # 确保 symbol 始终可从 state 还原（兼容 stock_code 旧键名）
                    if "symbol" not in last_state and "stock_code" in last_state:
                        last_state["symbol"] = last_state["stock_code"]

                # 根据节点类型构建特定消息与数据
                sse_event_type = "node_complete"
                extra_data     = {}
                human_msg      = f"✅ {label} 完成"

                if node_name == "data_node":
                    md = output.get("market_data", {})
                    human_msg  = (
                        f"数据获取完成: {symbol} | 最新价 {md.get('latest_price', 'N/A')} "
                        f"| 信号 {md.get('indicators', {}).get('tech_signal', 'N/A')}"
                    )
                    extra_data = {
                        "latest_price":    md.get("latest_price"),
                        "price_change_pct": md.get("price_change_pct"),
                        "tech_signal":      md.get("indicators", {}).get("tech_signal"),
                    }

                elif node_name == "rag_node":
                    rag = output.get("rag_context", "")
                    human_msg  = f"RAG 知识检索完成，获取 {len(rag)} 字符上下文"
                    extra_data = {"context_length": len(rag)}

                elif node_name == "fundamental_node":
                    report = output.get("fundamental_report", {}) or {}
                    human_msg  = (
                        f"基本面分析: {report.get('recommendation', 'N/A')} "
                        f"(置信度 {report.get('confidence', 0):.0%})"
                    )
                    extra_data = {
                        "recommendation": report.get("recommendation"),
                        "confidence":     report.get("confidence"),
                        "signal_strength": report.get("signal_strength"),
                        "reasoning_preview": (report.get("reasoning", "")[:100] + "..."),
                    }

                elif node_name == "technical_node":
                    report = output.get("technical_report", {}) or {}
                    human_msg  = (
                        f"技术分析: {report.get('recommendation', 'N/A')} "
                        f"(置信度 {report.get('confidence', 0):.0%})"
                    )
                    extra_data = {
                        "recommendation": report.get("recommendation"),
                        "confidence":     report.get("confidence"),
                        "signal_strength": report.get("signal_strength"),
                    }

                elif node_name == "sentiment_node":
                    report = output.get("sentiment_report", {}) or {}
                    human_msg  = (
                        f"舆情分析: {report.get('recommendation', 'N/A')} "
                        f"(置信度 {report.get('confidence', 0):.0%})"
                    )
                    extra_data = {
                        "recommendation": report.get("recommendation"),
                        "confidence":     report.get("confidence"),
                    }

                elif node_name == "portfolio_node":
                    has_conflict = output.get("has_conflict", False)
                    if has_conflict:
                        sse_event_type = "conflict"
                        human_msg = "⚡ 检测到基本面与技术面意见冲突，启动多空辩论机制..."
                        extra_data = {"conflict": True}
                    else:
                        human_msg = "基金经理综合决策完成，提交风控审核..."
                        extra_data = {"conflict": False}

                elif node_name == "debate_node":
                    outcome    = output.get("debate_outcome", {}) or {}
                    rounds     = output.get("debate_rounds", 1)
                    sse_event_type = "debate"
                    human_msg  = (
                        f"⚖️ 辩论第{rounds}轮裁决: {outcome.get('resolved_recommendation', 'N/A')} "
                        f"(置信度 {outcome.get('confidence_after_debate', 0):.0%}) "
                        f"— {outcome.get('deciding_factor', '')[:60]}"
                    )
                    extra_data = {
                        "resolved_recommendation": outcome.get("resolved_recommendation"),
                        "confidence_after_debate":  outcome.get("confidence_after_debate"),
                        "deciding_factor":          outcome.get("deciding_factor"),
                        "debate_rounds":            rounds,
                    }

                elif node_name == "risk_node":
                    rd         = output.get("risk_decision", {}) or {}
                    retries    = output.get("risk_rejection_count", 0)
                    approval   = rd.get("approval_status", "N/A")
                    status_map = {"APPROVED": "✅", "CONDITIONAL": "⚠️", "REJECTED": "❌"}
                    sse_event_type = "risk_check"

                    if approval == "REJECTED":
                        sse_event_type = "risk_retry"
                        human_msg = (
                            f"❌ 风控拒绝（第{retries}次）: {rd.get('rejection_reason', '')} "
                            f"→ 要求基金经理修订方案"
                        )
                    else:
                        human_msg = (
                            f"{status_map.get(approval, '?')} 风控审批: {approval} "
                            f"| 风险 {rd.get('risk_level')} "
                            f"| 仓位 {rd.get('position_pct', 0):.0f}%"
                        )
                    extra_data = {
                        "approval_status": approval,
                        "risk_level":      rd.get("risk_level"),
                        "position_pct":    rd.get("position_pct"),
                        "stop_loss_pct":   rd.get("stop_loss_pct"),
                        "take_profit_pct": rd.get("take_profit_pct"),
                        "rejection_reason": rd.get("rejection_reason"),
                    }

                elif node_name == "trade_executor":
                    order          = output.get("trade_order", {}) or {}
                    sse_event_type = "trade_order"
                    human_msg      = (
                        f"🎯 交易指令: {order.get('action', 'N/A')} {symbol} "
                        f"| 仓位 {order.get('quantity_pct', 0):.0f}% "
                        f"| 止损 {order.get('stop_loss')} "
                        f"| 止盈 {order.get('take_profit')}"
                    )
                    extra_data = order

                seq += 1
                yield _make_sse_event(
                    event=sse_event_type,
                    node=node_name,
                    message=human_msg,
                    data=extra_data,
                    seq=seq,
                )
                await asyncio.sleep(0)

    except BaseException as e:
        # 安全地序列化异常信息：str(KeyError('error')) = "'error'"，
        # 使用 type(e).__name__ + str(e) 组合避免误导性单引号输出
        _err_type = type(e).__name__
        _err_msg  = str(e) or repr(e)
        logger.error(
            f"[stream] 图执行异常 [{_err_type}]: {_err_msg}",
            exc_info=True,
        )
        _graph_error = f"{_err_type}: {_err_msg}"
        seq += 1
        yield _make_sse_event(
            event="error",
            node="system",
            message=f"分析过程异常 [{_err_type}]: {_err_msg}",
            data={"error": _err_msg, "error_type": _err_type},
            seq=seq,
        )
        # 不 return：继续构建降级完成事件，确保前端能渲染出部分结果

    # ── 发送完成事件（附最终 trade_order + 研报 + 图表数据）──────
    # complete 事件构建包裹在独立 try/except 内，避免 f-string None 值等次生错误断流
    seq += 1
    # 安全提取 symbol：优先用函数参数，再从 last_state 降级，最终回退 "UNKNOWN"
    symbol = symbol or last_state.get("symbol") or last_state.get("stock_code") or "UNKNOWN"
    try:
        final_order = last_state.get("trade_order") or {}

        # 编译 Markdown 深度研报（所有字段全部使用 .get() 安全提取）
        fundamental = last_state.get("fundamental_report") or {}
        technical   = last_state.get("technical_report")   or {}
        sentiment   = last_state.get("sentiment_report")   or {}
        debate      = last_state.get("debate_outcome")      or {}
        risk        = last_state.get("risk_decision")       or {}

        def _pct(v):
            try: return f"{float(v):.1%}"
            except: return str(v)

        _error_banner = (
            f"\n> ⚠ **分析流程异常中断**（{_graph_error}）。以下为各节点已完成的部分结果，供参考。\n\n---\n\n"
            if _graph_error else ""
        )
        markdown_report = f"""# {symbol} · 多智能体深度研报

> 生成时间：{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}　｜　分析引擎：LangGraph Multi-Agent
{_error_banner}
---

## 一、基本面分析

| 指标 | 结论 |
|------|------|
| 综合建议 | **{fundamental.get('recommendation', 'N/A')}** |
| 置信度   | {_pct(fundamental.get('confidence', 0))} |
| 信号强度 | {fundamental.get('signal_strength', 'N/A')} |

{fundamental.get('reasoning', '暂无详细推理') or '暂无详细推理'}

---

## 二、技术面分析

| 指标 | 结论 |
|------|------|
| 综合建议 | **{technical.get('recommendation', 'N/A')}** |
| 置信度   | {_pct(technical.get('confidence', 0))} |
| 信号强度 | {technical.get('signal_strength', 'N/A')} |

{technical.get('reasoning', '暂无详细推理') or '暂无详细推理'}

---

## 三、市场情绪分析

| 指标 | 结论 |
|------|------|
| 综合建议 | **{sentiment.get('recommendation', 'N/A')}** |
| 置信度   | {_pct(sentiment.get('confidence', 0))} |

{sentiment.get('reasoning', '暂无详细推理') or '暂无详细推理'}

---

## 四、风控决策

| 指标 | 数值 |
|------|------|
| 审批状态 | **{risk.get('approval_status', 'N/A')}** |
| 建议仓位 | {_pct(risk.get('position_pct') or 0)} |
| 止损线   | {_pct(risk.get('stop_loss_pct') or 0)} |
| 止盈线   | {_pct(risk.get('take_profit_pct') or 0)} |
| 风险等级 | {risk.get('risk_level', 'N/A')} |

{f"> 辩论裁决：{debate.get('resolved_recommendation','N/A')} | 决定因素：{debate.get('deciding_factor','')}" if debate else ""}

---

## 五、最终交易指令

| 字段 | 值 |
|------|----|
| 操作方向 | **{final_order.get('action', 'N/A')}** |
| 建议仓位 | {_pct(final_order.get('quantity_pct') or 0)} |
| 止损价   | {final_order.get('stop_loss', 'N/A')} |
| 止盈价   | {final_order.get('take_profit', 'N/A')} |
| 置信度   | {_pct(final_order.get('confidence', 0))} |

> **核心逻辑**：{final_order.get('rationale', '暂无')}

---

⚠ *本研报由 AI 多智能体自动生成，仅供学习参考，不构成投资建议。请遵守大学生守则：单股仓位 ≤ 15%，务必设置止损，切勿加杠杆。*
"""

        # 图表数据：从基本面 key_metrics 中提取实际历年数据
        key_metrics  = fundamental.get("key_metrics") or {}
        revenue_data = key_metrics.get("revenue_history") or []
        profit_data  = key_metrics.get("profit_history")  or []
        data_years   = key_metrics.get("years") or []
        # 优先使用真实历史年份；无数据时回退到近5年占位
        if data_years and revenue_data:
            chart_years  = [str(y) for y in data_years[-5:]]
            revenue_vals = [(v or 0) for v in revenue_data[-5:]]
            profit_vals  = [(v or 0) for v in profit_data[-5:]] if profit_data else [0] * len(chart_years)
            # 长度对齐
            n = len(chart_years)
            revenue_vals = (revenue_vals + [0] * n)[:n]
            profit_vals  = (profit_vals  + [0] * n)[:n]
        else:
            current_year = datetime.now(timezone.utc).year
            # 使用上一年往前推 4 年（当年财报通常未出齐）
            chart_years  = [str(current_year - 5 + i) for i in range(5)]
            revenue_vals = [0, 0, 0, 0, 0]
            profit_vals  = [0, 0, 0, 0, 0]
        # 若全为 0，提示前端"数据不足"
        has_chart_data = any(v != 0 for v in revenue_vals + profit_vals)

        # Phase 3 兜底：financial_chart_data / final_markdown_report 始终下发，前端无需防 undefined
        _chart_payload = {
            "years":         chart_years,
            "revenue":       revenue_vals,
            "profit":        profit_vals,
            "has_data":      has_chart_data,
            "revenue_label": key_metrics.get("revenue_label", "营业收入（亿元）"),
            "profit_label":  key_metrics.get("profit_label",  "净利润（亿元）"),
            "revenue_composition": key_metrics.get("revenue_composition", {}),
            "performance_trend":   key_metrics.get("performance_trend", {}),
        }
        yield _make_sse_event(
            event="complete",
            node="system",
            message=f"✅ 分析完成: {symbol} → {final_order.get('action', 'N/A')} "
                    f"(仓位 {float(final_order.get('quantity_pct') or 0):.0f}%)",
            data={
                "symbol":                symbol,
                "trade_order":           final_order,
                "status":                "completed",
                "final_markdown_report": markdown_report or "> 研报生成失败，请重试。",
                "financial_chart_data":  _chart_payload,
            },
            seq=seq,
        )
    except Exception as _complete_err:
        logger.error(
            f"[stream] complete 事件构建失败: {type(_complete_err).__name__}: {_complete_err}",
            exc_info=True,
        )
        yield _make_sse_event(
            event="complete",
            node="system",
            message=f"✅ 分析完成（部分数据渲染失败）: {symbol}",
            data={"symbol": symbol, "trade_order": {}, "status": "completed_partial",
                  "final_markdown_report": "> 研报渲染异常，请重试。", "financial_chart_data": {}},
            seq=seq,
        )


# ════════════════════════════════════════════════════════════════
# API 端点
# ════════════════════════════════════════════════════════════════

@app.post(
    "/api/v1/analyze",
    summary="流式分析交易标的",
    description=(
        "接收交易标的代码，启动 LangGraph 多智能体分析流程。\n"
        "通过 Server-Sent Events (SSE) 实时推送每个节点的分析进度与结果。\n\n"
        "客户端需使用 EventSource 或 httpx streaming 解析 SSE 事件流。"
    ),
    response_description="SSE 事件流（Content-Type: text/event-stream）",
)
async def analyze_symbol(request: AnalyzeRequest):
    """POST /api/v1/analyze — SSE 流式分析"""
    raw_symbol = request.symbol.strip()
    if not raw_symbol:
        raise HTTPException(status_code=400, detail="symbol 不能为空")

    # 模糊搜索拦截：中文/英文名称 → 标准交易代码
    symbol    = MarketClassifier.fuzzy_match(raw_symbol)
    thread_id = str(uuid.uuid4())

    logger.info(f"[API] 收到分析请求: 原始='{raw_symbol}' → 标准='{symbol}' | thread_id={thread_id}")

    return StreamingResponse(
        _stream_graph_events(symbol, thread_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control":     "no-cache",
            "X-Accel-Buffering": "no",    # Nginx 禁用缓冲
            "X-Thread-Id":       thread_id,
        },
    )


@app.get("/api/v1/health", response_model=HealthResponse)
async def health_check():
    """GET /api/v1/health — 健康检查"""
    from tools.knowledge_base import _ensemble_retriever

    return HealthResponse(
        status      = "ok",
        version     = "2.0.0",
        graph_ready = _compiled_graph is not None,
        kb_ready    = _ensemble_retriever is not None,
        timestamp   = datetime.now(timezone.utc).isoformat(),
    )


@app.get("/api/v1/graph/mermaid", summary="获取图拓扑 Mermaid 字符串")
async def get_graph_mermaid():
    """GET /api/v1/graph/mermaid — 返回 LangGraph 拓扑的 Mermaid 图字符串"""
    mermaid = """flowchart TD
    START([START]) --> data_node[数据情报员\\ndata_node]
    data_node --> fundamental_node[基本面分析师\\nfundamental_node]
    data_node --> technical_node[技术分析师\\ntechnical_node]
    data_node --> sentiment_node[舆情分析师\\nsentiment_node]
    data_node --> rag_node[RAG知识检索\\nrag_node]
    fundamental_node --> portfolio_node
    technical_node --> portfolio_node
    sentiment_node --> portfolio_node
    rag_node --> portfolio_node[基金经理\\nportfolio_node]
    portfolio_node -->|冲突检测| debate_node[多空辩论\\ndebate_node]
    portfolio_node -->|无冲突| risk_node[风控官\\nrisk_node]
    debate_node --> portfolio_node
    risk_node -->|REJECTED| portfolio_node
    risk_node -->|APPROVED| trade_executor[交易指令\\ntrade_executor]
    trade_executor --> END([END])"""

    if _compiled_graph:
        try:
            mermaid = _compiled_graph.get_graph().draw_mermaid()
        except Exception:
            pass  # 使用静态版本

    return {"mermaid": mermaid}


# ════════════════════════════════════════════════════════════════
# 持仓体检端点
# ════════════════════════════════════════════════════════════════

@app.post(
    "/api/v1/health-check",
    summary="持仓体检",
    description=(
        "接收持仓列表，启动 health_node 独立分支（START→health_node→END），\n"
        "返回持仓健康评分、集中度/回撤/流动性指标与 AI 优化建议。"
    ),
)
async def portfolio_health_check(request: PortfolioCheckRequest):
    """POST /api/v1/health-check — 持仓健康诊断（JSON 请求/响应）"""
    from graph.builder import build_health_graph, make_health_initial_state
    from graph.state import PortfolioPosition
    from utils.market_classifier import MarketClassifier

    # 构建 PortfolioPosition 列表
    positions = []
    for item in request.positions:
        market_type, _ = MarketClassifier.classify(item.symbol)
        positions.append(
            PortfolioPosition(
                symbol=item.symbol,
                market_type=market_type.value,
                quantity=item.quantity,
                avg_cost=item.avg_cost,
            ).model_dump()
        )

    if not positions:
        raise HTTPException(status_code=400, detail="持仓列表不能为空")

    # 运行持仓体检图（health_node 是 async def，必须用 ainvoke 而非 invoke）
    try:
        health_graph = build_health_graph()
        initial_state = make_health_initial_state(positions)
        result_state = await health_graph.ainvoke(initial_state)
    except Exception as e:
        logger.error(f"[health-check] 体检图执行失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"持仓体检执行失败: {str(e)}")

    health_report = result_state.get("health_report")
    if not health_report:
        raise HTTPException(status_code=500, detail="health_node 未返回体检报告，请检查 LLM 配置")

    return {
        "status": "ok",
        "health_report": health_report,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ════════════════════════════════════════════════════════════════
# 行情快讯 & 持仓摘要端点
# ════════════════════════════════════════════════════════════════

# 各市场默认观察股票列表
_MARKET_WATCHLISTS: dict[str, list[str]] = {
    "a":  ["600519.SH", "000858.SZ", "601318.SH", "002594.SZ",
           "300750.SZ", "600036.SH", "601899.SH", "000001.SZ"],
    "hk": ["00700.HK", "09988.HK", "03690.HK", "02318.HK", "01398.HK", "09999.HK"],
    "us": ["AAPL", "MSFT", "NVDA", "GOOGL", "AMZN", "TSLA", "META"],
}

# 后端内存模拟持仓（成本价基于某日收盘，演示用）
_MOCK_PORTFOLIO: list[dict] = [
    {"symbol": "600519.SH", "name": "贵州茅台", "quantity": 10,  "avg_cost": 1680.00},
    {"symbol": "000858.SZ", "name": "五粮液",   "quantity": 50,  "avg_cost":  148.00},
    {"symbol": "AAPL",      "name": "苹果",     "quantity": 20,  "avg_cost":  185.00},
    {"symbol": "00700.HK",  "name": "腾讯控股",  "quantity": 100, "avg_cost":  370.00},
]


# ════════════════════════════════════════════════════════════════
# 财商学长 AI 对话
# ════════════════════════════════════════════════════════════════

# 财商学长专属极速模型（独立于主链路，不影响 LangGraph 分析节点）
_CHAT_MODEL = "qwen3.5-flash"

_CAISHANG_SYSTEM_PROMPT = (
    "你现在是本量化沙盘系统的'财商学长'。\n"
    "你的角色：一个懂金融、精通量化交易的高年级学长。\n"
    "你的目标：用容易理解的语言解答同学们的金融疑问。\n\n"
    "【核心工作守则】\n"
    "1. 反应快速：回答控制在200字以内，直击要点，禁止废话铺垫。\n"
    "2. 语言风格：称呼对方'同学'，自称'学长'。用生活案例解释枯燥的金融术语。\n"
    "3. 合规红线：只做财商教育和逻辑科普，绝对禁止提供任何确定的买卖建议或荐股指令。"
    "遇到推荐股票的问题，请巧妙婉拒，并引导多看系统的基础面分析。\n"
    "4. 强调风险教育：提醒止损、仓位控制、不用生活费炒股等大学生守则。\n"
    "5. 禁止推荐任何具体真实券商或第三方平台。\n"
    "6. 非金融无关话题礼貌拒绝，说明只专注投资教育。\n"
    "7. 【严格边界】严禁强行分析实时行情或最新财报！"
    "当同学要求分析某只具体股票时，亲切引导：'同学，你可以在本平台的【个股分析】页面输入股票代码"
    "（如 00700.HK / 600519.SH / AAPL），召唤多智能体引擎获取基于实时数据的深度研报，那比学长靠谱多了！'"
)


@app.post("/api/v1/chat", summary="财商学长 AI 对话（Qwen + 持久记忆）")
async def chat_with_advisor(
    request: ChatRequest,
    db=Depends(_get_db_dep),
):
    """
    POST /api/v1/chat

    持久记忆逻辑:
      1. 前端在 localStorage 生成并持久保存一个 UUID 作为 session_key
      2. 每次对话携带 session_key → 后端从 DB 取最近 10 条消息作上下文
      3. 用户消息 + AI 回复均写入 chat_messages 表
      4. 无 session_key 时退化为无状态模式（兼容旧版 history 字段）
    """
    try:
        from utils.llm_client import LLMClient
        from db.crud import (
            get_or_create_chat_session,
            get_chat_history,
            append_chat_message,
            count_chat_messages,
        )
        loop = asyncio.get_event_loop()

        # ── 构建上下文 ─────────────────────────────────────────
        session_key   = request.session_key
        db_history    = []
        session_obj   = None

        if session_key:
            session_obj = await get_or_create_chat_session(db, session_key)
            db_messages = await get_chat_history(db, session_obj.id, limit=10)
            for m in db_messages:
                db_history.append({"role": m.role, "content": m.content})

        # 无 session_key 降级：使用客户端传来的 history（旧版兼容）
        context_turns = db_history if session_key else request.history[-6:]

        def _call_llm():
            client = LLMClient(provider="dashscope", model=_CHAT_MODEL)
            history_text = ""
            for turn in context_turns:
                role = "同学" if turn.get("role") == "user" else "学长"
                history_text += f"{role}：{turn.get('content', '')}\n"
            prompt = history_text + f"同学：{request.message}"
            return client.generate(
                prompt=prompt,
                system_prompt=_CAISHANG_SYSTEM_PROMPT,
                temperature=0.7,
                max_tokens=400,
            )

        reply = await loop.run_in_executor(None, _call_llm)

        # ── 持久化本轮对话 ──────────────────────────────────────
        if session_obj:
            await append_chat_message(db, session_obj.id, "user",      request.message)
            await append_chat_message(db, session_obj.id, "assistant", reply)

        return {
            "reply":       reply,
            "model":       _CHAT_MODEL,
            "session_key": session_key,
            "timestamp":   datetime.now(timezone.utc).isoformat(),
        }

    except Exception as e:
        logger.error(f"[chat] LLM 调用失败: {e}")
        raise HTTPException(status_code=502, detail=f"AI 对话服务暂时不可用: {str(e)}")


# ════════════════════════════════════════════════════════════════
# 用户认证端点
# ════════════════════════════════════════════════════════════════

class RegisterRequest(BaseModel):
    username: str = Field(..., min_length=2, max_length=50, description="???")
    email: str = Field(..., description="??")
    password: str = Field(..., min_length=6, description="??")


class LoginRequest(BaseModel):
    email: str = Field(..., description="??")
    password: Optional[str] = Field(default=None, description="??")


class AuthRegisterRequest(BaseModel):
    username: str = Field(..., min_length=2, max_length=50, description="???")
    email: str = Field(..., description="??")
    password: str = Field(..., min_length=6, description="??")
    verification_code: str = Field(..., min_length=6, max_length=6, description="?????")


class AuthLoginRequest(BaseModel):
    email: str = Field(..., description="??")
    password: Optional[str] = Field(default=None, description="??")
    verification_code: Optional[str] = Field(default=None, min_length=6, max_length=6, description="?????")


class SendCodeRequest(BaseModel):
    email: str = Field(..., description="??")
    purpose: str = Field(..., pattern="^(register|login)$", description="?????")


_CODE_TTL_MINUTES = 10
_CODE_COOLDOWN_SECONDS = 60


def _make_verification_code() -> str:
    return f"{random.randint(0, 999999):06d}"


@app.post("/api/v1/auth/send-code", summary="???????")
async def send_auth_code(request: SendCodeRequest, db=Depends(_get_db_dep)):
    from db.crud import get_email_verification_code, get_user_by_email, upsert_email_verification_code
    from utils.email_sender import is_email_configured, send_verification_email

    email = request.email.strip().lower()
    existing_user = await get_user_by_email(db, email)
    if request.purpose == "register" and existing_user:
        raise HTTPException(status_code=400, detail="???????")
    if request.purpose == "login" and not existing_user:
        raise HTTPException(status_code=404, detail="???????")
    if not is_email_configured():
        raise HTTPException(status_code=500, detail="???????????? SMTP")

    record = await get_email_verification_code(db, email, request.purpose)
    now = datetime.now(timezone.utc)
    if record and record.last_sent_at and (now - record.last_sent_at).total_seconds() < _CODE_COOLDOWN_SECONDS:
        raise HTTPException(status_code=429, detail="???????????????")

    code = _make_verification_code()
    await run_in_threadpool(send_verification_email, email, code, request.purpose)
    await upsert_email_verification_code(
        db, email, request.purpose, code, now + timedelta(minutes=_CODE_TTL_MINUTES)
    )
    return {
        "message": "????????????",
        "email": email,
        "purpose": request.purpose,
        "expires_in": _CODE_TTL_MINUTES * 60,
        "cooldown": _CODE_COOLDOWN_SECONDS,
    }


@app.post("/api/v1/auth/register", summary="???????")
async def register(request: AuthRegisterRequest, db=Depends(_get_db_dep)):
    from db.crud import (
        consume_email_verification_code,
        create_user,
        get_user_by_email,
        get_user_by_username,
    )
    from api.auth import create_access_token

    if await get_user_by_email(db, request.email):
        raise HTTPException(status_code=400, detail="???????")
    if await get_user_by_username(db, request.username):
        raise HTTPException(status_code=400, detail="????????")
    ok = await consume_email_verification_code(
        db,
        request.email.strip().lower(),
        "register",
        request.verification_code,
    )
    if not ok:
        raise HTTPException(status_code=400, detail="?????????")

    user = await create_user(db, request.username, request.email, request.password)
    token = create_access_token(user.id, user.username)
    return {
        "token": token,
        "user_id": user.id,
        "username": user.username,
        "message": "????????? CampusQuant?",
    }


@app.post("/api/v1/auth/login", summary="???????")
async def login(request: AuthLoginRequest, db=Depends(_get_db_dep)):
    from db.crud import consume_email_verification_code, get_user_by_email, verify_password
    from api.auth import create_access_token

    user = await get_user_by_email(db, request.email.strip().lower())
    if not user:
        raise HTTPException(status_code=401, detail="??????")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="??????")

    if request.verification_code:
        ok = await consume_email_verification_code(
            db,
            request.email.strip().lower(),
            "login",
            request.verification_code,
        )
        if not ok:
            raise HTTPException(status_code=401, detail="?????????")
    elif request.password:
        if not verify_password(request.password, user.hashed_password):
            raise HTTPException(status_code=401, detail="???????")
    else:
        raise HTTPException(status_code=400, detail="?????????")

    token = create_access_token(user.id, user.username)
    return {
        "token": token,
        "user_id": user.id,
        "username": user.username,
        "message": "????",
    }


from api.auth import get_current_user as _get_current_user, get_optional_user as _get_optional_user


@app.get("/api/v1/auth/me", summary="获取当前用户信息（需 Token）")
async def get_me(current_user=Depends(_get_current_user)):
    return {
        "user_id":    current_user.id,
        "username":   current_user.username,
        "email":      current_user.email,
        "bio":        current_user.bio,
        "avatar_url": current_user.avatar_url,
        "created_at": current_user.created_at.isoformat(),
    }


# ════════════════════════════════════════════════════════════════
# 模拟撮合下单
# ════════════════════════════════════════════════════════════════

# 市场短码 → VirtualAccount 资金字段名
_MT_CASH_FIELD = {"A": "cash_cnh", "HK": "cash_hkd", "US": "cash_usd", "UNKNOWN": "cash_cnh"}

# ── 港股每手股数字典（已知特殊手数，默认100） ────────────────────
# ── 港股手数动态缓存（HKEX 官方 ListOfSecurities.xlsx）──────────
import time as _time

_hk_lot_cache: dict[str, int] = {}   # "01211" -> 500 (5位无后缀)
_hk_lot_loaded_at: float = 0.0
_HK_LOT_CACHE_TTL: float = 86400.0   # 24小时刷新一次


async def _load_hk_lot_cache() -> None:
    """
    从 HKEX 官方 ListOfSecurities.xlsx 下载并解析 Board Lot 列，
    写入 _hk_lot_cache。失败时保留旧缓存并打 Warning。
    """
    global _hk_lot_cache, _hk_lot_loaded_at

    def _fetch() -> dict[str, int]:
        import io, requests
        from openpyxl import load_workbook
        url = ("https://www.hkex.com.hk/eng/services/trading/securities"
               "/securitieslists/ListOfSecurities.xlsx")
        hdrs = {"User-Agent": "Mozilla/5.0", "Referer": "https://www.hkex.com.hk"}
        r = requests.get(url, headers=hdrs, timeout=30)
        r.raise_for_status()
        # 注意: read_only=True 在部分 xlsx 中只读出极少行，必须用默认模式
        wb = load_workbook(io.BytesIO(r.content), read_only=False, data_only=True)
        ws = wb.active
        lot_map: dict[str, int] = {}
        for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
            code, lot = row[0], row[4]
            if not code or not lot:
                continue
            try:
                # Board Lot 可能以字符串 "1,000" 形式存储，去逗号后转 int
                lot_int = int(str(lot).replace(",", "").strip())
                lot_map[str(code).zfill(5)] = lot_int
            except (ValueError, TypeError):
                pass
        wb.close()
        return lot_map

    try:
        loop = asyncio.get_event_loop()
        lot_map = await loop.run_in_executor(None, _fetch)
        _hk_lot_cache = lot_map
        _hk_lot_loaded_at = _time.time()
        logger.info(f"✅ HKEX Board Lot 缓存加载完成（{len(lot_map)} 只港股）")
    except Exception as e:
        logger.warning(f"⚠️  HKEX Board Lot 缓存加载失败（fallback=100）: {e}")


async def get_min_lot(symbol: str, mt_short: str) -> int:
    """
    返回该标的最小交易手数（每手股数）。
    A股: 固定 100  |  US股: 固定 1
    港股: 实时查询 HKEX 官方 Board Lot 缓存，超时/失败降级为 100。
    """
    mt = mt_short.upper()
    if mt == "A":
        return 100
    if mt == "US":
        return 1
    # HK: 确保缓存有效（超过 TTL 则重新拉取）
    if _time.time() - _hk_lot_loaded_at > _HK_LOT_CACHE_TTL:
        await _load_hk_lot_cache()
    code = symbol.upper()
    if code.endswith(".HK"):
        code = code[:-3].zfill(5)
    lot = _hk_lot_cache.get(code)
    if lot is None:
        logger.debug(f"[lot] {symbol} 不在 HKEX 缓存中，使用默认 100")
        return 100
    return lot


async def _db_trade_order(db, user_id: int, symbol: str, action: str,
                           quantity: float, mt_short: str) -> dict:
    """
    已登录用户的 DB-backed 撮合引擎。
    直接从 DB 读取账户余额/持仓，执行撮合，写回 DB，不触碰全局内存账户。
    """
    from tools.market_data import get_spot_price_raw
    from db.crud import get_or_create_virtual_account, create_order as _db_order
    from db.models import VirtualAccount, Position
    from sqlalchemy import select

    cash_field = _MT_CASH_FIELD.get(mt_short.upper(), "cash_cnh")
    ts = datetime.now(timezone.utc).isoformat()

    # ── 数量基本校验 ────────────────────────────────────────────
    if quantity <= 0:
        return {"success": False, "error": "委托数量必须大于 0",
                "symbol": symbol, "action": action, "timestamp": ts}

    qty_int = int(quantity)
    if qty_int != quantity:
        return {"success": False, "error": "委托数量必须为整数（不支持小数股）",
                "symbol": symbol, "action": action, "timestamp": ts}

    # ── 手数（最小交易单位）校验（动态查询 HKEX）──────────────────
    min_lot = await get_min_lot(symbol, mt_short)
    if qty_int % min_lot != 0:
        mkt_label = {"A": "A股", "HK": "港股", "US": "美股"}.get(mt_short.upper(), "")
        return {"success": False,
                "error": f"{mkt_label} {symbol} 最小交易单位为 {min_lot} 股，"
                         f"请输入 {min_lot} 的整数倍（当前: {qty_int}）",
                "symbol": symbol, "action": action, "timestamp": ts}

    # 获取报价
    loop = asyncio.get_event_loop()
    try:
        spot = await loop.run_in_executor(None, get_spot_price_raw, symbol)
    except Exception:
        spot = {}
    exec_price = spot.get("price") or 0.0
    is_spot    = not spot.get("is_fallback", True)

    if not exec_price or exec_price <= 0:
        return {"success": False, "error": "行情获取失败，无法撮合",
                "symbol": symbol, "action": action, "timestamp": ts}

    amount = round(exec_price * quantity, 4)
    fee    = round(amount * 0.0003, 4)

    # 读取 DB 账户
    db_acct = await get_or_create_virtual_account(db, user_id)
    cash_before = getattr(db_acct, cash_field, 0.0)

    if action.upper() == "BUY":
        total_cost = amount + fee
        if total_cost > cash_before:
            return {"success": False,
                    "error": f"可用资金不足（需 {total_cost:.2f}，可用 {cash_before:.2f}）",
                    "symbol": symbol, "action": action, "timestamp": ts}
        cash_after = cash_before - total_cost
        setattr(db_acct, cash_field, cash_after)

        # Upsert position
        r = await db.execute(select(Position).where(
            Position.account_id == db_acct.id, Position.symbol == symbol))
        pos = r.scalar_one_or_none()
        stock_name = spot.get("name") or symbol
        if pos:
            total_qty    = pos.quantity + quantity
            pos.avg_cost = (pos.avg_cost * pos.quantity + exec_price * quantity) / total_qty
            pos.quantity = total_qty
            pos.market_type = mt_short
            # 若名称仍是裸代码则更新为真实名称
            if pos.name == symbol and stock_name != symbol:
                pos.name = stock_name
        else:
            db.add(Position(account_id=db_acct.id, symbol=symbol, name=stock_name,
                            quantity=quantity, avg_cost=exec_price, market_type=mt_short))

    else:  # SELL
        r = await db.execute(select(Position).where(
            Position.account_id == db_acct.id, Position.symbol == symbol))
        pos = r.scalar_one_or_none()
        held = pos.quantity if pos else 0.0
        if not pos or pos.quantity < quantity - 1e-8:
            return {"success": False,
                    "error": f"持仓不足（需 {quantity}，持有 {held:.2f}）",
                    "symbol": symbol, "action": action, "timestamp": ts}
        pos.quantity -= quantity
        if pos.quantity < 1e-6:
            await db.delete(pos)
        cash_after = cash_before + (amount - fee)
        setattr(db_acct, cash_field, cash_after)

    # 写成交记录（使用真实股票名称，不再裸用代码）
    _order_name = spot.get("name") or symbol
    await _db_order(db, db_acct.id, symbol=symbol, name=_order_name,
                    action=action.upper(), quantity=quantity, exec_price=exec_price,
                    amount=amount, fee=fee, cash_before=cash_before, cash_after=cash_after,
                    is_spot_price=is_spot, market_type=mt_short)
    await db.flush()

    return {
        "success":       True,
        "symbol":        symbol,
        "action":        action.upper(),
        "quantity":      quantity,
        "exec_price":    exec_price,
        "is_spot_price": is_spot,
        "amount":        amount,
        "fee":           fee,
        "cash_before":   cash_before,
        "cash_after":    cash_after,
        "simulated":     True,
        "timestamp":     ts,
    }


@app.post("/api/v1/trade/order", summary="模拟撮合下单（登录后持久化）")
async def place_trade_order(
    request: TradeOrderRequest,
    current_user=Depends(_get_optional_user),
    db=Depends(_get_db_dep),
):
    """
    POST /api/v1/trade/order

    - 未登录：使用内存全局账户（游客模式，重启清零）
    - 已登录：使用 DB 持久账户，订单写入数据库，重启不丢失
    """
    from utils.market_classifier import MarketClassifier

    symbol = request.symbol.strip().upper()
    symbol = MarketClassifier.fuzzy_match(symbol)
    market_type, _ = MarketClassifier.classify(symbol)
    mt_short = market_type.short   # "A" / "HK" / "US"

    # ── 手数和数量校验（对游客和登录用户均生效）─────────────────
    _qty = request.quantity
    _ts  = datetime.now(timezone.utc).isoformat()
    if _qty <= 0:
        return {"success": False, "error": "委托数量必须大于 0",
                "symbol": symbol, "action": request.action, "timestamp": _ts}
    _qty_int = int(_qty)
    if _qty_int != _qty:
        return {"success": False, "error": "委托数量必须为整数（不支持小数股）",
                "symbol": symbol, "action": request.action, "timestamp": _ts}
    _min_lot = await get_min_lot(symbol, mt_short)
    if _qty_int % _min_lot != 0:
        _mkt_label = {"A": "A股", "HK": "港股", "US": "美股"}.get(mt_short, "")
        return {"success": False,
                "error": f"{_mkt_label} {symbol} 最小交易单位为 {_min_lot} 股，"
                         f"请输入 {_min_lot} 的整数倍（当前: {_qty_int}）",
                "symbol": symbol, "action": request.action, "timestamp": _ts}

    # ── 已登录：直接从 DB 撮合，完全跳过全局内存账户 ─────────
    if current_user:
        return await _db_trade_order(
            db=db,
            user_id=current_user.id,
            symbol=symbol,
            action=request.action,
            quantity=request.quantity,
            mt_short=mt_short,
        )

    # ── 游客：使用内存全局账户（不持久化） ───────────────────
    from api.mock_exchange import get_account
    account = get_account()
    loop    = asyncio.get_event_loop()

    try:
        result = await loop.run_in_executor(
            None,
            lambda: account.place_order(
                symbol=symbol,
                action=request.action,
                quantity=request.quantity,
                market_type=mt_short,
            )
        )
    except Exception as e:
        logger.error(f"[trade/order] 撮合异常: {e}")
        raise HTTPException(status_code=502, detail=f"撮合引擎异常: {str(e)}")

    if not result.success:
        return {"success": False, "error": result.error, "symbol": symbol,
                "action": request.action, "timestamp": result.timestamp}

    return {
        "success":       True,
        "symbol":        result.symbol,
        "action":        result.action,
        "quantity":      result.quantity,
        "exec_price":    result.exec_price,
        "is_spot_price": result.is_spot_price,
        "amount":        result.amount,
        "fee":           result.fee,
        "cash_before":   result.cash_before,
        "cash_after":    result.cash_after,
        "simulated":     True,
        "timestamp":     result.timestamp,
    }


@app.get("/api/v1/market/search", summary="股票搜索联想（输入中英文/拼音/代码）")
async def search_stocks(q: str = ""):
    """
    GET /api/v1/market/search?q={query}

    对用户输入做多路模糊匹配，返回联想股票列表。
    优先查本地字典，再调新浪 Suggest API，总超时 3s。
    返回格式: { "suggestions": [{"symbol", "name", "type"}, ...] }
    """
    if not q or not q.strip():
        return {"suggestions": [], "query": q}

    try:
        loop = asyncio.get_event_loop()
        suggestions = await loop.run_in_executor(
            None,
            lambda: MarketClassifier.search_stock_suggestions(q.strip(), limit=8),
        )
    except Exception as e:
        logger.warning(f"[market/search] 搜索失败: {e}")
        suggestions = []

    return {"suggestions": suggestions, "query": q}


@app.get("/api/v1/market/quotes", summary="获取市场实时行情列表")
async def get_market_quotes(market: str = "a"):
    """
    GET /api/v1/market/quotes?market=a|hk|us

    调用 get_batch_quotes_raw() 获取指定市场的观察股票实时价格。
    返回标准化的报价列表，前端 market.html 使用。
    """
    market = market.lower()
    if market not in _MARKET_WATCHLISTS:
        raise HTTPException(status_code=400, detail=f"market 参数无效，可选: a / hk / us，收到: {market}")

    symbols = _MARKET_WATCHLISTS[market]

    try:
        from tools.market_data import get_batch_quotes_raw
        quotes = await run_in_threadpool(get_batch_quotes_raw, symbols, market)
    except Exception as e:
        logger.error(f"[market/quotes] 批量行情获取失败: {e}")
        raise HTTPException(status_code=502, detail=f"行情获取失败: {str(e)}")

    return {
        "market":    market,
        "quotes":    quotes,
        "count":     len(quotes),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@app.get("/api/v1/market/spot", summary="获取单只标的实时现价")
async def get_single_spot(symbol: str):
    """
    GET /api/v1/market/spot?symbol=600519.SH

    调用 get_spot_price_raw() 获取单个标的实时价格，供 trade.html 行情查询使用。
    返回: { symbol, name, price, change_pct, is_fallback, source }
    """
    if not symbol:
        raise HTTPException(status_code=400, detail="symbol 参数不能为空")
    try:
        from tools.market_data import get_spot_price_raw
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, get_spot_price_raw, symbol.upper())
        if not result or not result.get("price"):
            raise HTTPException(status_code=404, detail=f"未能获取 {symbol} 的行情数据")
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[market/spot] 单只行情获取失败 {symbol}: {e}")
        raise HTTPException(status_code=502, detail=f"行情获取失败: {str(e)}")


@app.get("/api/v1/market/lot/{symbol}", summary="获取标的最小交易手数（动态查询）")
async def get_lot_size(symbol: str):
    """
    GET /api/v1/market/lot/01211.HK  → {"symbol":"01211.HK","min_lot":500,"source":"hkex","market":"HK"}
    GET /api/v1/market/lot/600519.SH → {"symbol":"600519.SH","min_lot":100,"source":"fixed","market":"A"}
    GET /api/v1/market/lot/AAPL      → {"symbol":"AAPL","min_lot":1,"source":"fixed","market":"US"}

    港股从 HKEX 官方 Board Lot 缓存动态读取；A股固定 100；美股固定 1。
    """
    from utils.market_classifier import MarketClassifier
    symbol = symbol.strip().upper()
    symbol = MarketClassifier.fuzzy_match(symbol)
    market_type, _ = MarketClassifier.classify(symbol)
    mt_short = market_type.short

    min_lot = await get_min_lot(symbol, mt_short)
    source = "fixed" if mt_short in ("A", "US") else (
        "hkex" if _hk_lot_cache else "fallback"
    )
    return {
        "symbol":  symbol,
        "min_lot": min_lot,
        "source":  source,
        "market":  mt_short,
    }


@app.get("/api/v1/portfolio/summary", summary="获取虚拟账户持仓摘要（含实时估值）")
async def get_portfolio_summary(
    current_user=Depends(_get_optional_user),
    db=Depends(_get_db_dep),
):
    from tools.market_data import get_spot_price_raw

    if current_user:
        from db.crud import get_or_create_virtual_account, get_positions as _db_get_pos, get_orders
        db_acct = await get_or_create_virtual_account(db, current_user.id)
        db_pos  = await _db_get_pos(db, db_acct.id)
        raw_positions = [
            {"symbol": p.symbol, "name": p.name, "quantity": p.quantity,
             "avg_cost": p.avg_cost, "market_type": p.market_type}
            for p in db_pos
        ]
        cash_cnh = db_acct.cash_cnh
        cash_hkd = db_acct.cash_hkd
        cash_usd = db_acct.cash_usd
        orders   = await get_orders(db, db_acct.id, limit=1000)
        order_count = len(orders)
    else:
        from api.mock_exchange import get_account
        snap = get_account().snapshot()
        raw_positions = snap["positions"]
        cash_cnh = snap.get("cash_cnh", 0.0)
        cash_hkd = snap.get("cash_hkd", 0.0)
        cash_usd = snap.get("cash_usd", 0.0)
        order_count = int(snap.get("order_count") or 0)

    async def enrich_position(pos: dict) -> dict:
        loop = asyncio.get_event_loop()
        try:
            spot = await loop.run_in_executor(None, get_spot_price_raw, pos["symbol"])
        except Exception:
            spot = {"price": pos["avg_cost"], "change_pct": 0.0, "is_fallback": True}
        current_price  = spot.get("price") or pos["avg_cost"]
        change_pct     = spot.get("change_pct") or 0.0
        cost_value     = pos["quantity"] * pos["avg_cost"]
        market_value   = pos["quantity"] * current_price
        unrealized_pnl = market_value - cost_value
        pnl_pct        = (unrealized_pnl / cost_value * 100) if cost_value else 0.0
        return {
            **pos,
            "current_price":  round(current_price, 3),
            "change_pct":     round(change_pct, 2),
            "cost_value":     round(cost_value, 2),
            "market_value":   round(market_value, 2),
            "unrealized_pnl": round(unrealized_pnl, 2),
            "pnl_pct":        round(pnl_pct, 2),
            "is_fallback":    spot.get("is_fallback", False),
        }

    try:
        enriched = await asyncio.gather(*[enrich_position(p) for p in raw_positions])
    except Exception as e:
        logger.error(f"[portfolio/summary] 持仓摘要计算失败: {e}")
        raise HTTPException(status_code=502, detail=f"持仓摘要获取失败: {str(e)}")

    total_cost   = float(sum(p["cost_value"]     for p in enriched) or 0.0)
    total_market = float(sum(p["market_value"]   for p in enriched) or 0.0)
    total_pnl    = float(sum(p["unrealized_pnl"] for p in enriched) or 0.0)
    total_pnl_pct = (total_pnl / total_cost * 100) if total_cost else 0.0
    today_pnl    = float(sum(p["market_value"] * p["change_pct"] / 100 for p in enriched) or 0.0)
    total_assets = round(cash_cnh + total_market, 2)

    return {
        "cash":          round(cash_cnh, 2),
        "cash_cnh":      round(cash_cnh, 2),
        "cash_hkd":      round(cash_hkd, 2),
        "cash_usd":      round(cash_usd, 2),
        "total_assets":  total_assets,
        "positions":     list(enriched),
        "total_cost":    round(total_cost, 2),
        "total_market":  round(total_market, 2),
        "total_pnl":     round(total_pnl, 2),
        "total_pnl_pct": round(total_pnl_pct, 2),
        "today_pnl":     round(today_pnl, 2),
        "order_count":   order_count,
        "timestamp":     datetime.now(timezone.utc).isoformat(),
    }


# ════════════════════════════════════════════════════════════════
# 大盘指数 & 市场快讯端点
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/market/indices", summary="获取主要大盘指数实时行情")
async def get_market_indices():
    """
    GET /api/v1/market/indices

    优先返回后台缓存（<5ms）。缓存为空（刚启动瞬间）时触发一次紧急同步抓取。
    后台轮询每 5 分钟自动刷新，正常请求响应 <100ms。
    """
    if not _MARKET_CACHE["indices"]:
        # 刚启动缓存未就绪，触发紧急抓取（最多等 25s）
        logger.info("[market/indices] 缓存未就绪，触发紧急抓取")
        try:
            from tools.market_data import get_market_indices_raw
            loop = asyncio.get_event_loop()
            indices = await asyncio.wait_for(
                loop.run_in_executor(None, get_market_indices_raw), timeout=25.0
            )
            _MARKET_CACHE["indices"] = indices
        except Exception as e:
            logger.error(f"[market/indices] 紧急抓取失败: {e}")
            raise HTTPException(status_code=502, detail=f"大盘指数获取失败: {str(e)}")

    return {
        "indices":   _MARKET_CACHE["indices"],
        "count":     len(_MARKET_CACHE["indices"]),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@app.get("/api/v1/market/news", summary="获取财联社全球财经快讯")
async def get_market_news(limit: int = 20):
    """
    GET /api/v1/market/news?limit=20

    优先返回后台缓存（<5ms）。缓存为空（刚启动瞬间）时触发一次紧急同步抓取。
    后台轮询每 5 分钟自动刷新，正常请求响应 <100ms。
    """
    limit = max(1, min(limit, 50))
    if not _MARKET_CACHE["news"]:
        logger.info("[market/news] 缓存未就绪，触发紧急抓取")
        try:
            from tools.market_data import get_market_news_raw
            loop = asyncio.get_event_loop()
            news = await asyncio.wait_for(
                loop.run_in_executor(None, get_market_news_raw, 20), timeout=15.0
            )
            _MARKET_CACHE["news"] = news
        except Exception as e:
            logger.error(f"[market/news] 紧急抓取失败: {e}")
            raise HTTPException(status_code=502, detail=f"快讯获取失败: {str(e)}")

    news = _MARKET_CACHE["news"][:limit]
    return {
        "news":      news,
        "count":     len(news),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }



@app.get("/api/v1/market/sectors", summary="获取A股行业板块涨跌热力图数据")
async def get_market_sectors():
    """
    GET /api/v1/market/sectors

    返回约49个A股行业板块的实时涨跌幅，按涨跌幅降序排列。
    用于 market.html 右侧板块热力图展示。
    优先返回后台缓存（<5ms），缓存为空时触发紧急抓取。
    """
    if not _MARKET_CACHE["sectors"]:
        logger.info("[market/sectors] 缓存未就绪，触发紧急抓取")
        try:
            from tools.market_data import get_sector_data_raw
            loop = asyncio.get_event_loop()
            sectors = await asyncio.wait_for(
                loop.run_in_executor(None, get_sector_data_raw), timeout=12.0
            )
            _MARKET_CACHE["sectors"] = sectors
        except Exception as e:
            logger.error(f"[market/sectors] 紧急抓取失败: {e}")
            raise HTTPException(status_code=502, detail=f"板块数据获取失败: {str(e)}")

    return {
        "sectors":   _MARKET_CACHE["sectors"],
        "count":     len(_MARKET_CACHE["sectors"]),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@app.get("/api/v1/market/sentiment", summary="获取A股市场情绪实时指标")
async def get_market_sentiment():
    """
    GET /api/v1/market/sentiment

    返回四项实时指标：涨停家数、跌停家数、沪深成交额、北向资金净流入。
    优先返回后台缓存（<5ms），缓存为空时触发一次紧急同步抓取。
    """
    if not _MARKET_CACHE["sentiment"]:
        logger.info("[market/sentiment] 缓存未就绪，触发紧急抓取")
        try:
            from tools.market_data import get_market_sentiment_raw
            loop = asyncio.get_event_loop()
            sentiment = await asyncio.wait_for(
                loop.run_in_executor(None, get_market_sentiment_raw), timeout=20.0
            )
            _MARKET_CACHE["sentiment"] = sentiment
        except Exception as e:
            logger.error(f"[market/sentiment] 紧急抓取失败: {e}")
            raise HTTPException(status_code=502, detail=f"市场情绪数据获取失败: {str(e)}")

    return {
        **_MARKET_CACHE["sentiment"],
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ════════════════════════════════════════════════════════════════
# 成交记录（需登录）
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/trade/orders", summary="查询历史成交记录（需 Token）")
async def get_trade_orders(
    limit: int = 50,
    current_user=Depends(_get_current_user),
    db=Depends(_get_db_dep),
):
    from db.crud import get_or_create_virtual_account, get_orders

    db_account = await get_or_create_virtual_account(db, current_user.id)
    orders     = await get_orders(db, db_account.id, limit=min(limit, 100))

    return {
        "orders": [
            {
                "id":          o.id,
                "symbol":      o.symbol,
                "name":        o.name,
                "action":      o.action,
                "quantity":    o.quantity,
                "exec_price":  o.exec_price,
                "amount":      o.amount,
                "fee":         o.fee,
                "cash_after":  o.cash_after,
                "is_spot_price": o.is_spot_price,
                "market_type": o.market_type,
                "created_at":  o.created_at.isoformat(),
            }
            for o in orders
        ],
        "count":     len(orders),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ════════════════════════════════════════════════════════════════
# 投教社区端点
# ════════════════════════════════════════════════════════════════

class CreatePostRequest(BaseModel):
    title:   str = Field(..., min_length=5, max_length=200, description="帖子标题")
    content: str = Field(..., min_length=10, description="帖子正文")
    tag:     str = Field(default="learn", description="标签: learn|analysis|risk|exp")

    @field_validator("tag")
    @classmethod
    def validate_tag(cls, v: str) -> str:
        if v not in {"learn", "analysis", "risk", "exp"}:
            return "learn"
        return v


class CreateCommentRequest(BaseModel):
    content: str = Field(..., min_length=1, max_length=1000, description="评论内容")


def _format_post(post, username: str = "", liked: bool = False) -> dict:
    return {
        "id":          post.id,
        "title":       post.title,
        "content":     post.content,
        "tag":         post.tag,
        "like_count":  post.like_count,
        "view_count":  post.view_count,
        "author":      username or f"user_{post.user_id}",
        "user_id":     post.user_id,
        "liked":       liked,
        "created_at":  post.created_at.isoformat(),
        "updated_at":  post.updated_at.isoformat() if post.updated_at else None,
        # 摘要（前120字）
        "excerpt":     post.content[:120] + "…" if len(post.content) > 120 else post.content,
        "comment_count": 0,   # 由调用方覆盖（避免触发异步环境中的 lazy-load）
    }


@app.get("/api/v1/community/posts", summary="获取社区帖子列表")
async def list_posts(
    sort:   str = "latest",
    tag:    str = "",
    limit:  int = 20,
    offset: int = 0,
    current_user=Depends(_get_optional_user),
    db=Depends(_get_db_dep),
):
    from db.crud import get_posts, has_liked
    from db.crud import get_user_by_id as _get_user

    posts = await get_posts(
        db,
        tag_filter=tag if tag else None,
        sort=sort,
        limit=min(limit, 50),
        offset=offset,
    )

    from sqlalchemy import func as _func, select as _select
    from db.models import CommunityComment as _CC

    result = []
    for post in posts:
        author_user = await _get_user(db, post.user_id)
        liked = False
        if current_user:
            liked = await has_liked(db, current_user.id, post.id)
        # 查评论数（避免触发 lazy-load）
        cnt_r = await db.execute(_select(_func.count()).where(_CC.post_id == post.id))
        comment_count = cnt_r.scalar_one() or 0
        fmt = _format_post(post, username=author_user.username if author_user else "", liked=liked)
        fmt["comment_count"] = comment_count
        result.append(fmt)

    return {"posts": result, "count": len(result), "timestamp": datetime.now(timezone.utc).isoformat()}


@app.post("/api/v1/community/posts", summary="发布帖子（需 Token）")
async def create_post(
    request: CreatePostRequest,
    current_user=Depends(_get_current_user),
    db=Depends(_get_db_dep),
):
    from db.crud import create_post as db_create_post

    post = await db_create_post(db, current_user.id, request.title, request.content, request.tag)
    return {
        "message": "发帖成功",
        "post": _format_post(post, username=current_user.username),
    }


@app.get("/api/v1/community/posts/{post_id}", summary="获取帖子详情 + 评论")
async def get_post_detail(
    post_id: int,
    current_user=Depends(_get_optional_user),
    db=Depends(_get_db_dep),
):
    from db.crud import get_post, get_comments, has_liked
    from db.crud import get_user_by_id as _get_user

    post = await get_post(db, post_id)
    if not post:
        raise HTTPException(status_code=404, detail="帖子不存在")

    # 浏览量 +1
    post.view_count += 1

    author_user = await _get_user(db, post.user_id)
    liked = False
    if current_user:
        liked = await has_liked(db, current_user.id, post_id)

    comments_raw = await get_comments(db, post_id)
    comments = []
    for c in comments_raw:
        cu = await _get_user(db, c.user_id)
        comments.append({
            "id":         c.id,
            "content":    c.content,
            "author":     cu.username if cu else f"user_{c.user_id}",
            "created_at": c.created_at.isoformat(),
        })

    post_data = _format_post(post, username=author_user.username if author_user else "", liked=liked)
    post_data["comment_count"] = len(comments)
    return {"post": post_data, "comments": comments}


@app.post("/api/v1/community/posts/{post_id}/comments", summary="发表评论（需 Token）")
async def add_comment(
    post_id: int,
    request: CreateCommentRequest,
    current_user=Depends(_get_current_user),
    db=Depends(_get_db_dep),
):
    from db.crud import get_post, create_comment

    post = await get_post(db, post_id)
    if not post:
        raise HTTPException(status_code=404, detail="帖子不存在")

    comment = await create_comment(db, post_id, current_user.id, request.content)
    return {
        "message": "评论成功",
        "comment": {
            "id":         comment.id,
            "content":    comment.content,
            "author":     current_user.username,
            "created_at": comment.created_at.isoformat(),
        },
    }


@app.post("/api/v1/community/posts/{post_id}/like", summary="点赞/取消赞（需 Token）")
async def toggle_post_like(
    post_id: int,
    current_user=Depends(_get_current_user),
    db=Depends(_get_db_dep),
):
    from db.crud import get_post, toggle_like

    post = await get_post(db, post_id)
    if not post:
        raise HTTPException(status_code=404, detail="帖子不存在")

    liked = await toggle_like(db, current_user.id, post_id)
    return {
        "liked":      liked,
        "like_count": post.like_count,
    }


# ════════════════════════════════════════════════════════════════
# V1.2 — 三币种账户摘要
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/trade/account", summary="获取三币种虚拟账户摘要（含实时估值）")
async def get_trade_account(
    market: str = "",
    current_user=Depends(_get_optional_user),
    db=Depends(_get_db_dep),
):
    from tools.market_data import get_spot_price_raw

    mt_filter = market.upper() if market else None

    # ── 构建 positions 列表（DB 或内存）──────────────────────
    if current_user:
        from db.crud import get_or_create_virtual_account, get_positions as _db_get_pos
        db_acct   = await get_or_create_virtual_account(db, current_user.id)
        db_pos    = await _db_get_pos(db, db_acct.id)
        raw_positions = [
            {"symbol": p.symbol, "name": p.name, "quantity": p.quantity,
             "avg_cost": p.avg_cost, "market_type": p.market_type}
            for p in db_pos
            if (not mt_filter or p.market_type.upper() == mt_filter)
        ]
        balances = {
            "cash_cnh": db_acct.cash_cnh, "cash_hkd": db_acct.cash_hkd, "cash_usd": db_acct.cash_usd,
            "init_cnh": db_acct.init_cnh, "init_hkd": db_acct.init_hkd, "init_usd": db_acct.init_usd,
        }
    else:
        from api.mock_exchange import get_account
        account   = get_account()
        snap      = account.snapshot(market_type=mt_filter)
        raw_positions = snap["positions"]
        balances  = {k: snap.get(k, 0.0) for k in
                     ("cash_cnh","cash_hkd","cash_usd","init_cnh","init_hkd","init_usd")}

    # ── 异步富化（添加实时行情） ─────────────────────────────
    async def _enrich(pos: dict) -> dict:
        loop = asyncio.get_event_loop()
        try:
            spot = await loop.run_in_executor(None, get_spot_price_raw, pos["symbol"])
        except Exception:
            spot = {"price": pos["avg_cost"], "change_pct": 0.0, "is_fallback": True}
        cur_price = spot.get("price") or pos["avg_cost"]
        cost_val  = pos["quantity"] * pos["avg_cost"]
        mkt_val   = pos["quantity"] * cur_price
        pnl       = mkt_val - cost_val
        pnl_pct   = (pnl / cost_val * 100) if cost_val else 0.0
        return {
            **pos,
            "current_price":  round(cur_price, 4),
            "change_pct":     round(spot.get("change_pct") or 0.0, 2),
            "cost_value":     round(cost_val, 2),
            "market_value":   round(mkt_val, 2),
            "unrealized_pnl": round(pnl, 2),
            "pnl_pct":        round(pnl_pct, 2),
        }

    enriched = list(await asyncio.gather(*[_enrich(p) for p in raw_positions]))

    # ── Build per-market order count ────────────────────────────
    order_counts: dict[str, int] = {}
    if current_user:
        from db.crud import get_orders_by_market as _db_orders_by_mkt
        for _mkt in ("A", "HK", "US"):
            _ords = await _db_orders_by_mkt(db, db_acct.id, market_type=_mkt, limit=10000)
            order_counts[_mkt] = len(_ords)
    else:
        from api.mock_exchange import get_account as _get_acct
        _snap = _get_acct().snapshot()
        for o in _snap.get("orders", []):
            _m = (o.get("market_type") or "UNKNOWN").upper()
            order_counts[_m] = order_counts.get(_m, 0) + 1

    def _sub_account(currency: str, mkt: str) -> dict:
        cash_avail = balances.get(f"cash_{currency.lower()}", 0.0)
        init_val   = balances.get(f"init_{currency.lower()}", 0.0)
        sub_pos    = [p for p in enriched if p.get("market_type", "").upper() == mkt]
        mkt_val    = sum(p["market_value"]   for p in sub_pos)
        cost_val   = sum(p["cost_value"]     for p in sub_pos)
        pnl        = sum(p["unrealized_pnl"] for p in sub_pos)
        total      = cash_avail + mkt_val
        pnl_pct    = (pnl / cost_val * 100) if cost_val else 0.0
        total_pnl_abs = total - init_val
        total_pnl_pct = (total_pnl_abs / init_val * 100) if init_val else 0.0
        return {
            "market": mkt, "currency": currency,
            "cash": round(cash_avail, 2), "initial": round(init_val, 2),
            "market_value": round(mkt_val, 2), "total_assets": round(total, 2),
            "position_pnl": round(pnl, 2), "position_pnl_pct": round(pnl_pct, 2),
            "total_pnl": round(total_pnl_abs, 2), "total_pnl_pct": round(total_pnl_pct, 2),
            "positions": sub_pos,
            "order_count": order_counts.get(mkt, 0),
        }

    if mt_filter:
        _currency_map = {"A": "CNH", "HK": "HKD", "US": "USD"}
        cur = _currency_map.get(mt_filter, "CNH")
        return {"account": _sub_account(cur, mt_filter),
                "timestamp": datetime.now(timezone.utc).isoformat()}
    else:
        return {
            "accounts": {
                "A":  _sub_account("CNH", "A"),
                "HK": _sub_account("HKD", "HK"),
                "US": _sub_account("USD", "US"),
            },
            "positions_all": enriched,
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }


@app.get("/api/v1/trade/positions", summary="持仓列表（可按市场过滤）")
async def get_positions(
    market: str = "",
    current_user=Depends(_get_optional_user),
    db=Depends(_get_db_dep),
):
    """GET /api/v1/trade/positions?market=A|HK|US"""
    from tools.market_data import get_spot_price_raw

    mt_filter = market.upper() if market else None

    if current_user:
        from db.crud import get_or_create_virtual_account, get_positions as _db_get_pos
        db_acct = await get_or_create_virtual_account(db, current_user.id)
        db_pos  = await _db_get_pos(db, db_acct.id)
        raw_positions = [
            {"symbol": p.symbol, "name": p.name, "quantity": p.quantity,
             "avg_cost": p.avg_cost, "market_type": p.market_type}
            for p in db_pos
            if (not mt_filter or p.market_type.upper() == mt_filter)
        ]
    else:
        from api.mock_exchange import get_account
        snap = get_account().snapshot(market_type=mt_filter)
        raw_positions = snap["positions"]

    async def _enrich(pos: dict) -> dict:
        loop = asyncio.get_event_loop()
        try:
            spot = await loop.run_in_executor(None, get_spot_price_raw, pos["symbol"])
        except Exception:
            spot = {"price": pos["avg_cost"], "change_pct": 0.0, "is_fallback": True}
        cur_price = spot.get("price") or pos["avg_cost"]
        cost_val  = pos["quantity"] * pos["avg_cost"]
        mkt_val   = pos["quantity"] * cur_price
        pnl       = mkt_val - cost_val
        return {
            **pos,
            "current_price":  round(cur_price, 4),
            "change_pct":     round(spot.get("change_pct") or 0.0, 2),
            "cost_value":     round(cost_val, 2),
            "market_value":   round(mkt_val, 2),
            "unrealized_pnl": round(pnl, 2),
            "pnl_pct":        round((pnl / cost_val * 100) if cost_val else 0.0, 2),
        }

    enriched = list(await asyncio.gather(*[_enrich(p) for p in raw_positions]))
    return {
        "positions": enriched,
        "count":     len(enriched),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ════════════════════════════════════════════════════════════════
# V1.2 — K 线数据
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/market/kline", summary="获取 K 线数据（日/周/月）")
async def get_kline(
    symbol: str,
    period: str = "daily",    # daily | weekly | monthly
    count:  int  = 120,
):
    """
    GET /api/v1/market/kline?symbol=600519.SH&period=daily&count=120

    返回 Lightweight Charts candlestick 格式：
      [{"time":"2024-01-02","open":1.0,"high":1.2,"low":0.9,"close":1.1,"volume":1234}, ...]
    """
    symbol = symbol.strip().upper()
    symbol = MarketClassifier.fuzzy_match(symbol)
    period = period.lower()
    if period not in ("daily", "weekly", "monthly"):
        period = "daily"

    try:
        from tools.market_data import get_kline_data_raw
        loop  = asyncio.get_event_loop()
        kline = await loop.run_in_executor(
            None,
            lambda: get_kline_data_raw(symbol, period=period, count=min(count, 500))
        )
    except Exception as e:
        logger.error(f"[market/kline] {symbol} 获取失败: {e}")
        raise HTTPException(status_code=502, detail=f"K 线数据获取失败: {str(e)}")

    if not kline:
        raise HTTPException(status_code=404, detail=f"未找到 {symbol} 的 K 线数据")

    # 严格清洗：过滤 NaN / Inf / 零值 / 非法日期，保证 Lightweight Charts 不崩溃
    import math as _math, re as _re
    _OHLC = ("open", "high", "low", "close")
    _DATE_RE = _re.compile(r"^\d{4}-\d{2}-\d{2}$")

    def _is_valid(bar: dict) -> bool:
        # time 必须严格为 YYYY-MM-DD 格式
        t = bar.get("time", "")
        if not t or not _DATE_RE.match(str(t)):
            return False
        # 确保 time 字段是字符串（Lightweight Charts 要求）
        bar["time"] = str(t)[:10]
        for k in _OHLC:
            v = bar.get(k)
            if v is None:
                return False
            try:
                f = float(v)
                if _math.isnan(f) or _math.isinf(f) or f <= 0:
                    return False
                bar[k] = round(f, 4)   # 统一精度，消除浮点噪声
            except (TypeError, ValueError):
                return False
        # volume: NaN/null → 0，保留 0 volume（部分周期合法）
        vol = bar.get("volume")
        if vol is None or (isinstance(vol, float) and _math.isnan(vol)):
            bar["volume"] = 0
        return True

    kline_clean = [bar for bar in kline if _is_valid(bar)]
    if not kline_clean:
        raise HTTPException(status_code=404, detail=f"{symbol} K 线数据清洗后为空（全为 NaN/无效值）")

    # 按时间去重（相同 time 保留最后一条）并升序排列
    _seen: dict[str, dict] = {}
    for bar in kline_clean:
        _seen[bar["time"]] = bar
    kline_clean = sorted(_seen.values(), key=lambda b: b["time"])

    # 获取股票名称（best-effort，run_in_executor 避免阻塞 event loop）
    _name = symbol
    try:
        from tools.market_data import get_spot_price_raw
        _info = await loop.run_in_executor(None, get_spot_price_raw, symbol)
        _name = _info.get("name", symbol)
    except Exception:
        pass

    return {
        "symbol":    symbol,
        "name":      _name,
        "period":    period,
        "count":     len(kline_clean),
        "kline":     kline_clean,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ════════════════════════════════════════════════════════════════
# V1.2 — 多平台热榜
# ════════════════════════════════════════════════════════════════

# 后台定时刷新（15 min）
_hot_news_refresh_task: Optional[asyncio.Task] = None


@app.on_event("startup")
async def _start_hot_news_refresh():
    """启动后每 15 分钟后台刷新热榜缓存"""
    async def _loop():
        while True:
            try:
                from tools.hot_news import refresh_in_background
                refresh_in_background()
            except Exception as e:
                logger.warning(f"[hot_news] 定时刷新异常: {e}")
            await asyncio.sleep(15 * 60)

    global _hot_news_refresh_task
    _hot_news_refresh_task = asyncio.create_task(_loop())


_HOTNEWS_MOCK: list[dict] = [
    {
        "source": "cailian", "label": "财联社", "icon": "📰", "color": "#e74c3c",
        "items": [
            {"title": "沪深两市今日整体平稳，科技与消费板块轮番活跃", "url": "https://www.cls.cn/telegraph", "rank": 1},
            {"title": "央行最新数据显示社会融资规模稳步扩张", "url": "https://www.cls.cn/telegraph", "rank": 2},
            {"title": "多家券商发布年度策略报告，看好 A 股结构性机会", "url": "https://www.cls.cn/telegraph", "rank": 3},
        ], "fetched_at": None,
    },
    {
        "source": "xueqiu", "label": "雪球热搜", "icon": "❄️", "color": "#1db954",
        "items": [
            {"title": "贵州茅台", "url": "https://xueqiu.com/S/SH600519", "rank": 1},
            {"title": "宁德时代", "url": "https://xueqiu.com/S/SZ300750", "rank": 2},
            {"title": "中国平安", "url": "https://xueqiu.com/S/SH601318", "rank": 3},
        ], "fetched_at": None,
    },
    {
        "source": "zhihu", "label": "知乎热榜", "icon": "💬", "color": "#0084ff",
        "items": [
            {"title": "普通大学生如何科学规划第一笔投资？", "url": "https://www.zhihu.com/hot", "rank": 1},
            {"title": "ETF 定投和主动基金，哪个更适合新手？", "url": "https://www.zhihu.com/hot", "rank": 2},
            {"title": "财务自由的门槛到底有多高？", "url": "https://www.zhihu.com/hot", "rank": 3},
        ], "fetched_at": None,
    },
    {
        "source": "thepaper", "label": "澎湃新闻", "icon": "📌", "color": "#2ecc71",
        "items": [
            {"title": "国家统计局发布最新宏观经济数据", "url": "https://www.thepaper.cn/", "rank": 1},
            {"title": "证监会出台多项举措优化市场环境", "url": "https://www.thepaper.cn/", "rank": 2},
            {"title": "人民币汇率保持基本稳定，外汇储备规模充裕", "url": "https://www.thepaper.cn/", "rank": 3},
        ], "fetched_at": None,
    },
]


@app.get("/api/v1/market/hotnews", summary="多平台热榜聚合（财联社/雪球/知乎/凤凰/澎湃）")
async def get_hot_news(force: bool = False):
    """
    GET /api/v1/market/hotnews?force=false

    返回 5 个平台各 Top 3 热门内容列表（直接返回 list，非 wrapper object）。
    force=true 立即重新抓取（仅调试用）。
    若全部来源均抓取失败，返回内置 Mock 数据，保证前端不空白。
    """
    try:
        from tools.hot_news import get_hot_news as _get_hot_news
        loop = asyncio.get_event_loop()
        data = await loop.run_in_executor(None, lambda: _get_hot_news(force_refresh=force))
    except Exception as e:
        logger.error(f"[market/hotnews] 获取失败，降级到 Mock: {e}")
        data = []

    # 兜底：若所有来源均为空，返回 Mock 数据
    if not data or all(not src.get("items") for src in data):
        logger.warning("[market/hotnews] 所有来源为空，返回 Mock 兜底数据")
        return _HOTNEWS_MOCK

    # 直接返回 list（前端期望 list，不需要 wrapper）
    return data


# ════════════════════════════════════════════════════════════════
# V1.2 — Dashboard 聚合接口
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/dashboard/summary", summary="Dashboard 聚合数据（账户摘要+快讯）")
async def get_dashboard_summary(
    current_user=Depends(_get_optional_user),
    db=Depends(_get_db_dep),
):
    from tools.market_data import get_market_news_raw

    # ── 账户余额（已登录从 DB 读取，游客读内存）──────────────
    if current_user:
        from db.crud import get_or_create_virtual_account
        db_acct = await get_or_create_virtual_account(db, current_user.id)
        account_overview = {
            "A":  {"market": "A",  "currency": "CNH", "cash": round(db_acct.cash_cnh, 2), "initial": round(db_acct.init_cnh, 2)},
            "HK": {"market": "HK", "currency": "HKD", "cash": round(db_acct.cash_hkd, 2), "initial": round(db_acct.init_hkd, 2)},
            "US": {"market": "US", "currency": "USD", "cash": round(db_acct.cash_usd, 2), "initial": round(db_acct.init_usd, 2)},
        }
    else:
        from api.mock_exchange import get_account
        snap = get_account().snapshot()
        account_overview = {
            "A":  {"market": "A",  "currency": "CNH", "cash": snap["cash_cnh"], "initial": snap["init_cnh"]},
            "HK": {"market": "HK", "currency": "HKD", "cash": snap["cash_hkd"], "initial": snap["init_hkd"]},
            "US": {"market": "US", "currency": "USD", "cash": snap["cash_usd"], "initial": snap["init_usd"]},
        }

    flash_news = []
    try:
        loop       = asyncio.get_event_loop()
        flash_news = await loop.run_in_executor(None, get_market_news_raw, 5)
    except Exception as e:
        logger.warning(f"[dashboard/summary] 快讯获取失败: {e}")

    user_info = None
    if current_user:
        user_info = {
            "user_id":    current_user.id,
            "username":   current_user.username,
            "avatar_url": current_user.avatar_url,
        }

    return {
        "account_overview": account_overview,
        "flash_news":       flash_news,
        "user_info":        user_info,
        "timestamp":        datetime.now(timezone.utc).isoformat(),
    }


# ════════════════════════════════════════════════════════════════
# 财商学长 AI 聊天端点
# ════════════════════════════════════════════════════════════════

class MentorChatRequest(BaseModel):
    message: str
    history: list[dict] = []

@app.post("/api/v1/chat/mentor", summary="财商学长 AI 对话")
async def chat_mentor(req: MentorChatRequest):
    """
    POST /api/v1/chat/mentor
    Body: { message: str, history: [{role, content}] }
    Returns: { reply: str }
    """
    try:
        from langchain_core.messages import HumanMessage, AIMessage, SystemMessage

        sys_prompt = (
            "你是「财商学长」，一位专注于大学生财商教育的 AI 助手。"
            "你的职责是帮助大学生理解金融基础知识，包括：股票投资入门、ETF定投策略、"
            "仓位风险管理、市盈率/市净率解读、防范投资骗局等。"
            "回答要简洁易懂，多用举例，避免专业术语堆砌。"
            "提醒用户本平台仅为模拟练习，不构成投资建议。"
            "每次回答控制在 200 字以内。"
        )

        messages = [SystemMessage(content=sys_prompt)]
        for h in req.history[-8:]:
            role = h.get("role", "user")
            content = h.get("content", "")
            if role == "user":
                messages.append(HumanMessage(content=content))
            elif role == "assistant":
                messages.append(AIMessage(content=content))
        messages.append(HumanMessage(content=req.message))

        from graph.nodes import _build_llm
        llm = _build_llm(temperature=0.7)
        resp = await asyncio.get_event_loop().run_in_executor(None, llm.invoke, messages)
        reply = resp.content if hasattr(resp, "content") else str(resp)
        return {"reply": reply}

    except Exception as e:
        logger.warning(f"[chat/mentor] LLM 调用失败: {e}")
        return {"reply": f"学长暂时离线了，请稍后再试。（{type(e).__name__}）"}


# ════════════════════════════════════════════════════════════════
# 开发模式入口
# ════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "api.server:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        log_level="info",
    )

# ── 静态文件托管（必须放在所有路由定义之后）──────────────────────
# 访问 http://127.0.0.1:8000/market.html 等同于直接打开 HTML 文件
# 但带有正确的 HTTP Origin，不会触发 CORS 拦截
import os as _os
_PROJECT_ROOT = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
app.mount("/", StaticFiles(directory=_PROJECT_ROOT, html=True), name="static")
